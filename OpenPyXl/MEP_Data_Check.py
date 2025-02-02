import pandas as pd
from openpyxl import workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from colorsys import hsv_to_rgb

def generate_color(n):
    colors = []
    for i in range(n):
        hue = i/n
        saturation = 0.5+(i%2)*0.25
        value = 0.8
        rgb = hsv_to_rgb(hue, saturation, value)
        hex_color = ''.join(f'{int(c*255):02X}' for c in rgb)
        colors.append(hex_color)
    return colors

input_file = 'MEP.xlsx'
output_file = 'MEP_Out.xlsx'

df = pd.read_excel(input_file)

unique_elements = df['IdentificationNumber'].drop_duplicates().tolist()

workbook = load_workbook(input_file)
sheet = workbook.active

colors = generate_color(len(unique_elements))
fills = [PatternFill(start_color=color, end_color=color, fill_type='solid') for color in colors]

# Create a mapping of unique elements to their fill styles
element_to_fill = {unique_element: fills[idx % len(fills)] for idx, unique_element in enumerate(unique_elements)}

# Track which unique elements have matches
matched_elements = set()

# Highlight matches in columns 3–6 and track matches
for row in range(2, sheet.max_row + 1):  # Assuming header is in row 1
    for col in range(4, 8):  # Columns 4 to 8
        cell = sheet.cell(row=row, column=col)
        if cell.value in element_to_fill:  # Match found
            matched_elements.add(cell.value)  # Add to matched elements
            cell.fill = element_to_fill[cell.value]

# Apply color to the 3rd column only if the element is in matched_elements
for row in range(2, sheet.max_row + 1):  # Assuming header is in row 1
    cell = sheet.cell(row=row, column=3)  # Column 2 (1-based index)
    if cell.value in matched_elements:  # Color only if there's a match in 3–6 columns
        cell.fill = element_to_fill[cell.value]
    else:
        cell.fill = PatternFill(fill_type=None)

unmatched_elements = [element for element in unique_elements if element not in matched_elements]

new_col_index = sheet.max_column+1
sheet.cell(row=1, column=new_col_index, value="Unmatched Elements")
for idx, element in enumerate(unmatched_elements, start=2):
    sheet.cell(row=idx, column=new_col_index, value = element)

workbook.save(output_file)
